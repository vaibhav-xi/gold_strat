import csv
from datetime import datetime, timedelta
from tabulate import tabulate
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

csv_file = "xau_usd_m15.csv"

def parse_time(timestr):
    return datetime.strptime(timestr.strip(), '%Y-%m-%d %H:%M:%S')

# Read and parse data
data = []
with open(csv_file, mode='r', newline='', encoding='utf-8') as file:
    reader = csv.reader(file, delimiter='\t')
    headers = next(reader, None)  # Skip header

    for row in reader:
        if len(row) < 6 or row[0].lower().strip() == 'time':
            continue
        try:
            data.append({
                'Time': parse_time(row[0]),
                'Open': float(row[1]),
                'High': float(row[2]),
                'Low': float(row[3]),
                'Close': float(row[4]),
                'Volume': int(row[5]),
                'Other': row[6] if len(row) > 6 else None
            })
        except Exception as e:
            print(f"Skipping row due to error: {e}")

# Group data by date
by_date = defaultdict(list)
for row in data:
    date_key = row['Time'].date()
    by_date[date_key].append(row)

results = []

for date, rows in by_date.items():
    # Sort data by time just in case
    rows.sort(key=lambda x: x['Time'])

    # Get the 13:30:00 candle
    base_candle = next((r for r in rows if r['Time'].time() == datetime.strptime("13:30:00", "%H:%M:%S").time()), None)
    if not base_candle:
        continue  # skip if no 13:30

    base_time = base_candle['Time']
    base_high = base_candle['High']
    base_low = base_candle['Low']
    base_close = base_candle['Close']
    
    if (base_high - base_low) < 4:
        continue

    # Get 13:45 candle for SL
    sl_candle = next((r for r in rows if r['Time'].time() == datetime.strptime("13:15:00", "%H:%M:%S").time()), None)
    if not sl_candle:
        continue

    # Find breakout
    bias = None
    breakout_candle = None
    for r in rows:
        if r['Time'] <= base_time + timedelta(minutes=15):
            continue  # skip until 13:45+
        if r['Close'] > base_high:
            bias = "Buy"
            breakout_candle = r
            break
        elif r['Close'] < base_low:
            bias = "Sell"
            breakout_candle = r
            break

    if not breakout_candle:
        continue  # No breakout, skip

    entry_price = breakout_candle['Close']
    sl_price = sl_candle['Low'] if bias == 'Buy' else sl_candle['High']
    breakout_time = breakout_candle['Time']
    
    sl_points = abs(entry_price - sl_price)

    # Track performance
    sl_hit = False
    max_points = 0

    for r in rows:
        if r['Time'] <= breakout_time:
            continue
        if bias == "Buy":
            if r['Low'] <= sl_price:
                sl_hit = True
                break
            move = r['High'] - entry_price
            max_points = max(max_points, move)
        elif bias == "Sell":
            if r['High'] >= sl_price:
                sl_hit = True
                break
            move = entry_price - r['Low']
            max_points = max(max_points, move)

    # Assess trade
    result = {
        "Date": date.strftime("%Y-%m-%d"),
        "13:30 High": base_high,
        "13:30 Low": base_low,
        "13:30 Close": base_close,
        "Breakout Time": breakout_time.strftime("%H:%M:%S"),
        "Breakout Close (Entry)": entry_price,
        "Bias": bias,
        "SL": sl_price,
        "SL Points": sl_points,
        "Max Points": round(max_points, 2),
        "SL Hit": "Yes" if sl_hit else "No",
        "Trade Result": "Bad" if sl_hit and max_points < 5 else "Good"
    }

    results.append(result) 

# Display table
if results:
    print(tabulate(results, headers="keys", tablefmt="pretty"))
else:
    print("No valid breakout trades found.")


wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Gold Strategy Backtest"

headers = list(results[0].keys())
ws.append(headers)

for result in results:
    ws.append(list(result.values()))

for col_num, column_title in enumerate(headers, 1):
    column_letter = get_column_letter(col_num)
    ws.column_dimensions[column_letter].width = max(len(str(column_title)) + 2, 15)

for cell in ws[1]:
    cell.font = Font(bold=True)

total_trades = len(results)
good_trades = sum(1 for r in results if r["Trade Result"] == "Good")
efficiency = round((good_trades / total_trades) * 100, 2) if total_trades else 0.0

avg_sl_points = round(sum(r["SL Points"] for r in results) / len(results), 2) if results else 0.0

ws.append([])
ws.append(["Efficiency Summary"])
ws.append(["Average SL Points", avg_sl_points])
ws.append(["Total Trades", total_trades])
ws.append(["Good Trades", good_trades])
ws.append(["Efficiency (%)", efficiency])

output_file = "gold_strategy_backtest.xlsx"
wb.save(output_file)

print(f"\nExcel file saved as: {output_file}")   